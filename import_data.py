#!/usr/bin/env python3
"""Import data asli dari Excel (MsKaryawan + TrGajiKaryawan) ke SQLite FastGaji."""
import sys, os, sqlite3
sys.path.insert(0, "/opt/data/fastgaji")
os.chdir("/opt/data/fastgaji")

import openpyxl
from datetime import datetime

# buat tabel dulu (pakai definisi dari fastgaji.py)
import fastgaji as fg
fg.init_db()

# ===== KONFIG =====
XLSX_KARYAWAN = "/opt/data/cache/documents/doc_958ee793f2a1_MsKaryawan.xlsx"
XLSX_HISTORI = "/opt/data/cache/documents/doc_5c5b9d664302_TrGajiKaryawan.xlsx"
DB_FILE = "/opt/data/fastgaji/fastgaji.db"

def _db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def _fmt_jatuh_tempo(dt):
    """datetime -> '11-Aug-25'"""
    if not dt:
        return ""
    try:
        months = ["Jan","Feb","Mar","Apr","Mei","Jun","Jul","Agu","Sep","Okt","Nov","Des"]
        return f"{dt.day}-{months[dt.month-1]}-{str(dt.year)[-2:]}"
    except:
        return str(dt)

def _clean(s):
    if s is None:
        return ""
    return str(s).replace("_x000d_", " ").replace("\r", " ").replace("\n", " ").strip()[:2000]

# ===== 1. IMPORT KARYAWAN =====
print("=== IMPORT MsKaryawan ===")
wb = openpyxl.load_workbook(XLSX_KARYAWAN, read_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
print(f"  {len(rows)-1} karyawan ditemukan")

conn = _db()
count = 0
for r in rows[1:]:
    if r[0] is None and r[1] is None:
        continue
    kode = int(r[0]) if r[0] is not None else 0
    if kode == 0:
        continue
    gaji = int(r[5]) if r[5] else 0
    jt = _fmt_jatuh_tempo(r[6])
    conn.execute("""INSERT OR REPLACE INTO karyawan
        (kode, nama, gaji, gaji_x_fee, eth, pintu, jatuh_tempo, komisi, notes,
         bank, norek, namarek, email, notelp, status, ewallet, btc, ovo)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (kode, _clean(r[1]), gaji, int(gaji * 1.002),
         _clean(r[15]) if len(r) > 15 else "", _clean(r[20]) if len(r) > 20 else "",
         jt, _clean(r[10]) if len(r) > 10 else "", _clean(r[9]) if len(r) > 9 else "",
         _clean(r[2]) if len(r) > 2 else "", _clean(r[3]) if len(r) > 3 else "",
         _clean(r[4]) if len(r) > 4 else "", _clean(r[7]) if len(r) > 7 else "",
         _clean(r[8]) if len(r) > 8 else "", r[11] if len(r) > 11 and r[11] is not None else 0,
         _clean(r[13]) if len(r) > 13 else "", _clean(r[14]) if len(r) > 14 else "",
         _clean(r[19]) if len(r) > 19 else ""))
    count += 1
conn.commit()
print(f"  ✅ {count} karyawan di-import!")

# ===== 2. IMPORT HISTORI GAJI =====
print("=== IMPORT TrGajiKaryawan ===")
wb2 = openpyxl.load_workbook(XLSX_HISTORI, read_only=True)
ws2 = wb2[wb2.sheetnames[0]]
rows2 = list(ws2.iter_rows(values_only=True))
print(f"  {len(rows2)-1} transaksi gaji ditemukan")

conn.execute("DELETE FROM histori")
hcount = 0
for r in rows2[1:]:
    if r[1] is None:
        continue
    kode = int(r[1]) if r[1] else 0
    gaji = int(r[2]) if r[2] else 0
    tgl = ""
    if r[3]:
        try:
            tgl = r[3].strftime("%d-%b-%y")
        except:
            tgl = str(r[3])[:16]
    notes = _clean(r[4]) if len(r) > 4 else ""
    conn.execute("INSERT INTO histori (kode_karyawan, nama, gaji_idr, tanggal, notes, waktu) VALUES (?,?,?,?,?,?)",
                 (kode, "", gaji, tgl, notes, tgl))
    hcount += 1
conn.commit()
conn.close()
print(f"  ✅ {hcount} transaksi histori di-import!")

# ===== 3. VERIFIKASI =====
conn = _db()
k = conn.execute("SELECT COUNT(*) as c FROM karyawan").fetchone()["c"]
h = conn.execute("SELECT COUNT(*) as c FROM histori").fetchone()["c"]
print(f"\n=== FINAL ===")
print(f"  Karyawan: {k}")
print(f"  Histori: {h}")
# sample
sample = conn.execute("SELECT kode, nama, gaji, eth, pintu, jatuh_tempo FROM karyawan LIMIT 5").fetchall()
for s in sample:
    print(f"    {s['kode']} | {s['nama'][:30]} | {s['gaji']:,} | {s['eth'][:14]}...")
conn.close()
print("\n✅ IMPORT SELESAI — database FastGaji siap!")
